import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

data = json.load(open('/tmp/final_results_500.json'))
data.sort(key=lambda x: x['code'])

FONT_NAME = 'Arial'

wb = Workbook()
ws = wb.active
ws.title = '資料'

headers = [
    ('code', '代號', 10, '@'),
    ('name', '名稱', 14, None),
    ('industry', '產業分類', 14, None),
    ('q1_gm', 'Q1毛利率%', 12, '0.00'),
    ('q1_om', 'Q1營業利益率%', 12, '0.00'),
    ('q1_nm', 'Q1稅後純益率%', 12, '0.00'),
    ('q2_gm', 'Q2毛利率%', 12, '0.00'),
    ('q2_om', 'Q2營業利益率%', 12, '0.00'),
    ('q2_nm', 'Q2稅後純益率%', 12, '0.00'),
    ('g_gm', '毛利率成長(pp)', 13, '+0.00;-0.00'),
    ('g_om', '營業利益率成長(pp)', 14, '+0.00;-0.00'),
    ('g_nm', '稅後純益率成長(pp)', 14, '+0.00;-0.00'),
    ('q1_eps', 'Q1 EPS', 10, '0.00'),
    ('h1_eps', 'H1 EPS', 10, '0.00'),
    ('q2_eps', 'Q2 EPS', 10, '0.00'),
    ('per', '目前本益比', 10, '0.00'),
    ('price', '目前股價', 10, '0.00'),
    ('dividend', '今年配息(元)', 12, '0.00'),
    ('yield', '殖利率%', 10, '0.00'),
]

n_cols = len(headers)
n_rows = len(data)

header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
header_font = Font(name=FONT_NAME, bold=True, color='FFFFFF', size=10)
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin = Side(style='thin', color='D0D0D0')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for c, (key, label, width, fmt) in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=c, value=label)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align
    cell.border = border
    ws.column_dimensions[get_column_letter(c)].width = width

ws.row_dimensions[1].height = 30

col_idx = {key: i + 1 for i, (key, *_rest) in enumerate(headers)}

data_font = Font(name=FONT_NAME, size=10)
name_align = Alignment(horizontal='left')
num_align = Alignment(horizontal='right')

for r, row in enumerate(data, start=2):
    for key, label, width, fmt in headers:
        c = col_idx[key]
        cell = ws.cell(row=r, column=c)
        cell.font = data_font
        cell.border = border
        if key in ('g_gm', 'g_om', 'g_nm', 'yield'):
            continue
        val = row.get(key)
        if key == 'code':
            cell.value = val
            cell.alignment = name_align
        elif key in ('name', 'industry'):
            cell.value = val if val is not None else ''
            cell.alignment = name_align
        else:
            cell.value = val
            cell.alignment = num_align
            if fmt:
                cell.number_format = fmt

    q1gm_l = get_column_letter(col_idx['q1_gm'])
    q1om_l = get_column_letter(col_idx['q1_om'])
    q1nm_l = get_column_letter(col_idx['q1_nm'])
    q2gm_l = get_column_letter(col_idx['q2_gm'])
    q2om_l = get_column_letter(col_idx['q2_om'])
    q2nm_l = get_column_letter(col_idx['q2_nm'])

    c = ws.cell(row=r, column=col_idx['g_gm'])
    c.value = f'=IF(OR({q1gm_l}{r}="",{q2gm_l}{r}=""),"",{q2gm_l}{r}-{q1gm_l}{r})'
    c.font = data_font; c.border = border; c.alignment = num_align; c.number_format = '+0.00;-0.00'

    c = ws.cell(row=r, column=col_idx['g_om'])
    c.value = f'=IF(OR({q1om_l}{r}="",{q2om_l}{r}=""),"",{q2om_l}{r}-{q1om_l}{r})'
    c.font = data_font; c.border = border; c.alignment = num_align; c.number_format = '+0.00;-0.00'

    c = ws.cell(row=r, column=col_idx['g_nm'])
    c.value = f'=IF(OR({q1nm_l}{r}="",{q2nm_l}{r}=""),"",{q2nm_l}{r}-{q1nm_l}{r})'
    c.font = data_font; c.border = border; c.alignment = num_align; c.number_format = '+0.00;-0.00'

    price_l = get_column_letter(col_idx['price'])
    div_l = get_column_letter(col_idx['dividend'])
    yield_c = ws.cell(row=r, column=col_idx['yield'])
    yield_c.value = f'=IF(OR({price_l}{r}="",{div_l}{r}="",{price_l}{r}=0),"",{div_l}{r}/{price_l}{r}*100)'
    yield_c.font = data_font; yield_c.border = border; yield_c.alignment = num_align; yield_c.number_format = '0.00'

last_row = n_rows + 1
last_col_letter = get_column_letter(n_cols)

ws.freeze_panes = 'C2'
ws.auto_filter.ref = f'A1:{last_col_letter}{last_row}'

green_font = Font(name=FONT_NAME, size=10, color='1E7B34')
red_font = Font(name=FONT_NAME, size=10, color='C0392B')

pos_neg_cols = ['q1_gm', 'q1_om', 'q1_nm', 'q2_gm', 'q2_om', 'q2_nm',
                'g_gm', 'g_om', 'g_nm', 'q1_eps', 'h1_eps', 'q2_eps']
for key in pos_neg_cols:
    col_l = get_column_letter(col_idx[key])
    rng = f'{col_l}2:{col_l}{last_row}'
    ws.conditional_formatting.add(rng, CellIsRule(operator='greaterThan', formula=['0'], font=red_font))
    ws.conditional_formatting.add(rng, CellIsRule(operator='lessThan', formula=['0'], font=green_font))

for r in range(2, last_row + 1):
    ws.row_dimensions[r].height = 16

ws2 = wb.create_sheet('說明')
ws2.column_dimensions['A'].width = 100
title_font = Font(name=FONT_NAME, bold=True, size=13)
label_font = Font(name=FONT_NAME, bold=True, size=10.5)
body_font = Font(name=FONT_NAME, size=10.5)
wrap = Alignment(wrap_text=True, vertical='top')

notes = [
    ('title', '台股前500大上市櫃公司 2026 Q1 vs Q2 財報三率、EPS、本益比、殖利率比較 — 方法論說明'),
    ('gap', ''),
    ('label', '範圍'),
    ('body', '台灣上市＋上櫃依市值排名前500大公司（排除金融控股／銀行／證券／保險業）。資料日期約 2026/08/21。'),
    ('gap', ''),
    ('label', '三率／EPS 資料來源'),
    ('body', '取自 FinMind TaiwanStockFinancialStatements，2026年Q1／Q2單季數字直接計算三率；成長率(pp)＝Q2三率−Q1三率，試算表中以公式自動計算。'),
    ('gap', ''),
    ('label', '產業分類／本益比'),
    ('body', '取自 FinMind TaiwanStockInfo／TaiwanStockPER。虧損或無意義者留白。'),
    ('gap', ''),
    ('label', '目前股價'),
    ('body', '取自 FinMind TaiwanStockPrice 最新收盤價；每日排程自動更新。'),
    ('gap', ''),
    ('label', '今年配息／殖利率'),
    ('body', '配息＝FinMind TaiwanStockDividend加總除息日在2026年內之現金股利；殖利率＝配息÷股價×100%，試算表中以公式自動計算，非官方年化數值。'),
    ('gap', ''),
    ('label', '⚠️ 資料完整度提醒'),
    ('body', '新納入約245家公司中，因API速率限制（每欄位最多重試3次），約150家的三率／EPS／本益比／配息暫缺（留白），僅代號、名稱、產業、股價較完整。原有255家資料完整，之後可再補齊。'),
    ('gap', ''),
    ('label', '欄位性質'),
    ('body', '「成長(pp)」與「殖利率%」為試算表公式，隨左側資料自動重算；其餘欄位皆為外部資料匯入之數值。'),
    ('gap', ''),
    ('label', '免責聲明'),
    ('body', '此為第三方彙整資料，僅供參考，正式數字請以公司公告及證交所／櫃買中心資訊為準。'),
]

r = 1
for kind, text in notes:
    cell = ws2.cell(row=r, column=1, value=text)
    if kind == 'title':
        cell.font = title_font
    elif kind == 'label':
        cell.font = label_font
    elif kind == 'body':
        cell.font = body_font
        cell.alignment = wrap
        ws2.row_dimensions[r].height = 45
    r += 1

wb.save('/tmp/site_build/taiwan_stocks_top500.xlsx')
print('Saved. Rows:', n_rows, 'Cols:', n_cols)
